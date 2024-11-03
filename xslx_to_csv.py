import openpyxl
import csv

def xlsx_to_csv(xlsx_file, csv_file):
    workbook = openpyxl.load_workbook(xlsx_file)
    worksheet = workbook.active  # Используем активный лист
    
    with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Запись заголовков
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            row_data = [cell.value for cell in row]
            writer.writerow(row_data)
        
        # Запись данных
        for row in worksheet.iter_rows(min_row=2):
            row_data = [cell.value for cell in row]
            writer.writerow(row_data)

# Пример использования
xlsx_file = 'siriys_tablica111.xlsx'  # Замените на ваш файл
csv_file = 'siriys_tablica111.csv'  # Замените на имя вашего CSV-файла
xlsx_to_csv(xlsx_file, csv_file)